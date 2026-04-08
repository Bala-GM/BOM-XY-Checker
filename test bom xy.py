import pandas as pd
import numpy as np
import os
from os import getcwd
import csv as csv
from io import StringIO
import re
import datetime
from io import BytesIO
import time
import sys
from datetime import datetime
import shutil
import linecache
import ast
import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox 
from tkinter.filedialog import askopenfile
import subprocess
import threading
import urllib
import urllib.parse
from sqlite3 import dbapi2 as sqlite
import sqlite3
import lxml
import openpyxl
from openpyxl.workbook import workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlsxwriter
import seaborn as ssn
import matplotlib
import matplotlib as mlp
from matplotlib import pyplot as plt
import xlwings as xw
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pyxcelframe
from openpyxl import load_workbook
from pyxcelframe import copy_cell_style
from pyxcelframe import insert_frame
from pyxcelframe import insert_columns
from pyxcelframe import sheet_to_sheet
from pyxcelframe import column_last_row
#import PySimpleGUI as sg
import sqlalchemy
from sqlalchemy import create_engine
import sqlalchemy_access as sq_a
import sqlalchemy_pyodbc_mssql as sqlalchemy
from flask_sqlalchemy import SQLAlchemy
import pyodbc
import pypyodbc
import odbc
import psycopg2
import mysql.connector as sql
import MySQLdb as sql #pip install mysqlclient
from plyer import notification
import xlrd
import xml.etree.ElementTree as ET
from openpyxl.utils.exceptions import InvalidFileException
import psutil
import hashlib
import wmi

#########################################################################################################################################################################
print('\n')
print("\033[92;4m*******XY Data Manipulation*******\033[0m")
print('\n')
#########################################################################################################################################################################

try:
    # XY Data MANIPULATION
    os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM')
    file_path_xlsx = 'BOM.xlsx'
    file_path_xls = 'BOM.xls'

    if os.path.isfile(file_path_xlsx):
        dsxy1 = pd.read_excel(file_path_xlsx, sheet_name="XY DATA", usecols=['R', 'X', 'Y', 'A', 'S'], index_col=False)
    elif os.path.isfile(file_path_xls):
        dsxy1 = pd.read_excel(file_path_xls, sheet_name="XY DATA", usecols=['R', 'X', 'Y', 'A', 'S'], index_col=False)
    else:
        root = tk.Tk()
        root.withdraw()
        error_message = "BOM file not found. Please check the file path."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    dfXY1 = dsxy1

    # Define your column lists
    column_list_1 = ['R', 'X', 'Y', 'A', 'S']

    # Check which column list is present in the DataFrame
    if all(column in dsxy1.columns for column in column_list_1):
        columns_to_use = column_list_1
    else:
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        missing_columns = [column for column in column_list_1 if column not in dsxy1.columns]
        error_message = f"The following columns are missing: {', '.join(missing_columns)}"
        error_msgbm1 = "The following columns are missing in BOM (EXCEL) File Sheet 'XY DATA':\n'Reference CRD as R'\n'X Coordinate as X'\n'Y Coordinate as Y'\n'Angle as A'\n'Side as S'"
        messagebox.showerror("Error", error_message)
        messagebox.showerror("Error", error_msgbm1)
        sys.exit(1)  # Exit the program with an error code

    dfXY1['R'] = dfXY1['R'].str.replace(" ","")
    dfXY1.rename(columns={'R': 'B_Ref_List'}, inplace=True)
    dfXY1['R'] = dfXY1['B_Ref_List']

    dfXYC1 = len(dfXY1['B_Ref_List'])
    print(f'Total count of rows in the "XY_Ref_List" column: {dfXYC1}') # TAKING XY REFERENCE COUNT FROM THE BEIGN
    dfXYC2 = dfXY1['S'].value_counts().reset_index()
    
    # Find duplicates based on X and Y
    duplicates = dfXY1[dfXY1.duplicated(subset=["X", "Y"], keep=False)]

    # Sort them
    duplicates_sorted = duplicates.sort_values(by=["X", "Y"])

    print(duplicates_sorted)
    
    duplicates_sorted["COUNT"] = duplicates_sorted.groupby(["X", "Y"])["R"].transform("count")
    
    df_duplicates = duplicates_sorted[duplicates_sorted["COUNT"] > 1]
    
    df_duplicates = df_duplicates.sort_values(by=["X", "Y"])

    #dsplcr1['Long Des'] = dsplcr1['Long Des'].str.replace('0201', '')
    #dsplcr1['Long Des'] = dsplcr1['Long Des'].str.replace('0402', '')
    #dsplcr1['Long Des'] = dsplcr1['Long Des'].str.replace('0603', '')
    #dsplcr1['Long Des'] = dsplcr1['Long Des'].str.replace('0805', '')
    #dsplcr1['Long Des'] = dsplcr1['Long Des'].str.replace('1206', '')

except Exception as e:
    # Handle the exception gracefully
    error_message = f"An error occurred: {e}"

    # Show error message in a pop-up box
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    messagebox.showerror("Error", error_message)
    error_msgbm1 = f"The following columns are missing in BOM Sheet 'XY DATA':\n'Reference CRD as R'\n'X Coordinate as X'\n'Y Coordinate as Y'\n'Angle as A'\n'Side as S'"
    messagebox.showerror("Error", error_msgbm1)
    sys.exit(1)  # Exit the program with an error code


with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/BOM_List_OP.xlsx") as writer:

    #ds1.to_excel(writer, sheet_name="BOM", index=False)
    dfXY1.to_excel(writer, sheet_name="XY DATA", index=False)
    #dsn2.to_excel(writer, sheet_name="AVL GROUP", index=False)
    #dcn1.to_excel(writer, sheet_name="PART MASTER", index=False)
    #df_AL1.to_excel(writer, sheet_name="AVL SHEET", index=True)
    #ds3.to_excel(writer, sheet_name="BOM DATA", index=False)
    #count_df = pd.DataFrame({'BOM Ref Count': [dsco3]})
    #count_df.to_excel(writer, sheet_name="BOM Ref Count", index=TRUE)
    count_df = pd.DataFrame({'XY Ref Count': [dfXYC1]})
    duplicates_sorted.to_excel(writer, sheet_name="XY duplicates_sorted", index=False)
    df_duplicates.to_excel(writer, sheet_name="XY duplicates_sort", index=False)
    count_df.to_excel(writer, sheet_name="XY Ref Count", index=TRUE)
    dfXYC2.to_excel(writer, sheet_name="XY Side Counts", index=True)
    #dsplcr1.to_excel(writer, sheet_name="BOM_SL", index=False)
    #dnsplcr1.to_excel(writer, sheet_name="Orginal_BOM_SL", index=False)

pass
print('The file does not exist.')
#########################################################################################################################################################################
#########################################################################################################################################################################
